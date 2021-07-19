import pandas as pd
from openpyxl import Workbook
import openpyxl

class Excel:

#Création d'une liste pour stocker les élémenents du fichier xlsx
    def __init__(self):
        self.dico = {}          # Dictionnaire afin de stoker les strings et de les convertire en int
        self.compteur = 0       # Compteur
        self.row = []           # Ligne courrante
        self.new_file = []      # Liste pour stocker les nouvelles lignes

#Change les élements de la liste liste en entier
    def change_to_int_list(self,liste,c):
        new_liste =[]
        for i in liste:
            if c > 0:
                if i not in self.dico :
                    self.dico[i] = self.compteur
                    self.compteur+=1
                new_liste.append(self.dico[i])
            else : 
                new_liste.append(i)
        self.new_file.append(new_liste)

#Récupération de tout les elments du fichier xlsx
    def read_clean_file(self,name,maxi=-1):
        print('Reading file ...')
        book = openpyxl.load_workbook(name,read_only=True)
        sheet = book.active
        book_clean = Workbook()
        sheet_clean = book_clean.active
        print('File readed')
        c = 0
        for value in sheet.iter_rows(min_row=1, min_col=1, values_only=True): 
            self.row.append(value)
            self.change_to_int_list(self.row[0],c)
            sheet_clean.append(self.new_file[0])
            self.row = []
            self.new_file = []
            c+=1
        print('end')
        book_clean.save("final.xlsx")